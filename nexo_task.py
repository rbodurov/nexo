import requests

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

try:
    # set boolean for futher logic
    skip = False
    addresses = []
    my_api_key = "5QYME84DIWV3CRUZGVIM6CZDSUSRCT688Z"
    # testing addresses
    # 0xeC20607aa654D823DD01BEB8780a44863c57Ed07
    # 0xfF1b44f1FCCebc4890B5E00a1EA9259d00a40fEb

    print("How many addresses do you want to retrieve information for?")
    print("For one address select - 1")
    print("For more than one address select - 2")

    while not skip:
        try:
            answer = input("Select answer: ")
            if int(answer) == 1:
                print("Enter an address: ")
                address = input()
                addresses.append(address)
                skip = True
            elif int(answer) == 2:
                print("Enter addresses (separated by commas): ")
                input_addresses = input()
                addresses = [x.strip() for x in input_addresses.split(',')]
                skip = True
            else:
                print("Wrong answer! Choices are 1 and 2.")
        except ValueError:
            print("You've entered an incorrect answer!")

    final_result = []

    for address in addresses:
        # retrieve balance data
        api_url_balance = "https://api.polygonscan.com/api?module=account&action=balance&address={}&apikey={}".format(address, my_api_key)
        balance_response = requests.get(api_url_balance)
        balance_data = balance_response.json()
        balance_data = int(balance_data['result']) / 10 ** 18

        # retrieve transactions data
        api_url_transactions = "https://api.polygonscan.com/api?module=account&action=txlist&address={}&startblock=0&endblock=99999999&sort=desc&apikey={}".format(address, my_api_key)
        transactions_response = requests.get(api_url_transactions)
        transactions_data = transactions_response.json()

        # gather and format all needed data
        user_data = {}
        user_data['address'] = address
        user_data['balance'] = balance_data
        user_data['transactions'] = transactions_data['result']
        final_result.append(user_data)

    # create table
    wb = Workbook()
    ws = wb.active

    for item in final_result:
        headings = ['Address', 'Balance']
        ws.append(headings)

        data_row = []
        data_row.append('{}'.format(item['address']))
        data_row.append('{} MATIC'.format(round(item['balance'], 3)))
        ws.append(data_row)

        txn_headings = ['Txn Hash', 'Block', 'Date Time (UTC)', 'From', 'To', 'Value']
        ws.append(txn_headings)

        for txn in item['transactions']:
            txn_row = []
            txn_row.append('%s' % txn['hash'])
            txn_row.append('%s' % txn['blockNumber'])
            txn_row.append('%s' % datetime.fromtimestamp(int(txn['timeStamp'])))
            txn_row.append('%s' % txn['from'])
            txn_row.append('%s' % txn['to'])
            txn_row.append('%s' % int(int(txn['value']) / 10 ** 18))
            ws.append(txn_row)

        empty_row = []
        ws.append(empty_row)

        header_cells = headings + txn_headings

        # table styles
        this_border = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=1, min_col=1):
            for cell in row:
                if cell.value:
                    cell.border = Border(top=this_border, left=this_border, right=this_border, bottom=this_border)
                    if cell.value in header_cells:
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(fgColor="00C0C0C0", fill_type="solid")

        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20

    wb.save("user_data.xlsx")
    print("The information has been downloaded and saved in your current directory!")
except Exception as e:
    print("Oops!", e.__class__, "occurred.")
